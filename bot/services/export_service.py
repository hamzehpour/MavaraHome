"""Builds an in-memory .xlsx export of all reservations for the admin panel."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from database.repositories import reservations as reservations_repo
from texts import fa
from utils.jalali import gregorian_iso_to_jalali_display


def export_reservations_xlsx() -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reservations"
    ws.sheet_view.rightToLeft = True

    headers = [
        "کد رزرو", "نام خریدار", "شماره خریدار", "نام حاضر", "شماره حاضر",
        "تعداد", "مبلغ کل", "وضعیت", "تاریخ سانس", "ساعت سانس", "تاریخ ثبت",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="2F4858", end_color="2F4858", fill_type="solid")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in reservations_repo.list_all():
        session_date_fa = gregorian_iso_to_jalali_display(r["session_date"]) if r.get("session_date") else "-"
        ws.append([
            r.get("reservation_code") or "-",
            r.get("user_full_name"),
            r.get("user_phone"),
            r.get("attendee_name") or "",
            r.get("attendee_phone") or "",
            r.get("people"),
            r.get("total_price"),
            fa.STATUS_LABELS.get(r.get("status"), r.get("status")),
            session_date_fa,
            r.get("session_time") or "-",
            r.get("created_at"),
        ])

    widths = [14, 22, 14, 22, 14, 8, 12, 16, 22, 10, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    buffer.name = "reservations_export.xlsx"
    return buffer


def export_logs_xlsx() -> io.BytesIO:
    """
    Full audit trail — every reservation/session/event create/edit/delete,
    approvals, rejections, capacity changes, staff changes, etc. — so the
    admin always has something to check "چه کسی چه چیزی را چه‌وقت تغییر داد".
    """
    from database.repositories import logs as logs_repo

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Log"
    ws.sheet_view.rightToLeft = True

    headers = ["زمان", "اکشن", "آیدی تلگرام انجام‌دهنده", "جزئیات"]
    ws.append(headers)
    header_fill = PatternFill(start_color="2F4858", end_color="2F4858", fill_type="solid")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for entry in logs_repo.list_all():
        ws.append([
            entry.get("created_at"),
            entry.get("action"),
            entry.get("telegram_id") or "-",
            entry.get("details") or "",
        ])

    widths = [20, 28, 20, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    buffer.name = "audit_log_export.xlsx"
    return buffer


def export_reservations_csv() -> io.BytesIO:
    import csv
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([
        "reservation_code", "buyer_name", "buyer_phone", "attendee_name", "attendee_phone",
        "people", "total_price", "status", "session_date", "session_time", "created_at",
    ])
    for r in reservations_repo.list_all():
        writer.writerow([
            r.get("reservation_code") or "", r.get("user_full_name"), r.get("user_phone"),
            r.get("attendee_name") or "", r.get("attendee_phone") or "",
            r.get("people"), r.get("total_price"), r.get("status"),
            r.get("session_date"), r.get("session_time"), r.get("created_at"),
        ])
    byte_buffer = io.BytesIO(buffer.getvalue().encode("utf-8-sig"))  # BOM so Excel opens Persian text correctly
    byte_buffer.name = "reservations_export.csv"
    return byte_buffer


def export_reservations_json() -> io.BytesIO:
    import json
    data = reservations_repo.list_all()
    byte_buffer = io.BytesIO(json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8"))
    byte_buffer.name = "reservations_export.json"
    return byte_buffer
